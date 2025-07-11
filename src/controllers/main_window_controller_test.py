import logging
import math
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import numpy as np
from typing import Callable, Literal
import os
import sys
sys.path.insert(0,os.path.join(os.path.abspath(os.path.dirname(__file__)),"..",".."))
from src.modules.AnomalyDetection import AnomalyDetection
from src.modules.Utils import (
    quick_stability_check,
    calculate_focus_score,
    get_resource_path,
    open_file,
    draw_focus_score,
)
from PySide6.QtWidgets import QFileDialog
import cv2
import requests
from PySide6.QtCore import (
    QObject,
    QSize,
    Qt,
    QThread,
    QTimer,
    Signal,
    Slot,
)

from PySide6.QtGui import (
    QIcon,
    QImage,
    QPixmap,
)

from PySide6.QtWidgets import QApplication, QLabel, QMainWindow, QMessageBox

from src.global_params import (
    camera_config_params,
    abnormal_inference_params,
    sfc_request_params,
    product_config_params,
    detect_lcd_params,
    save_all_config,
    system_config_params,
)

from src.views.main_window_test import Ui_MainWindow
from src.modules.Loggers import console_logger, operation_history_log

from PySide6.QtCore import QObject, QEvent


class ResizeFilter(QObject):
    def __init__(self, label):
        super().__init__()
        self.label = label

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.Resize:
            size = self.label.size()
            print(f"Kích thước QLabel: {size.width()} x {size.height()}")
        return super().eventFilter(obj, event)

class StreamVideoWorker(QObject):
    emit_frame_signal = Signal(object)
    finished_signal = Signal()

    def __init__(self, parent):
        super().__init__()
        self._parent: MainWindowController = parent
        self.focus_score: int = 0
        self.count_fr: int = 0
        self.cap: cv2.VideoCapture | None = None
        self.is_running: bool = False
        self.abnormal_processing: AnomalyDetection | None = None

        self.init_video_capture()
        # camera

        try:
            self.abnormal_processing = AnomalyDetection()
        except Exception as e:
            console_logger.error(f"Error initializing StreamVideoWorker: {e}")
            operation_history_log.error(f"Error initializing StreamVideoWorker: {e}")
        if self.abnormal_processing is None:
            console_logger.error("Abnormal processing is None")

    def rotate_image(self, image: np.ndarray, angle: float) -> np.ndarray:
        """
        Rotates an image by a specified angle around its center.
        Args:
            image (np.ndarray): The input image to be rotated.
            angle (float): The angle (in degrees) to rotate the image. Positive values mean counter-clockwise rotation.
        Returns:
            np.ndarray: The rotated image with the same dimensions as the input.
        """
        
        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)  # Tâm xoay là trung tâm ảnh

        # Tạo ma trận xoay 2D
        M = cv2.getRotationMatrix2D(center, angle, scale=1.0)

        # Xoay ảnh
        rotated = cv2.warpAffine(image, M, (w, h))
        return rotated

    def init_video_capture(self) -> None:
        """
        Initializes the video capture device with the specified camera configuration parameters.
        This method releases any previously opened video capture device, then attempts to open a new
        video capture using the source and settings defined in `camera_config_params`. It supports
        both DirectShow and default capture backends, and can enable MJPG format if specified.
        Camera properties such as brightness, contrast, saturation, sharpness, gain, resolution,
        auto exposure, and exposure are set according to the configuration.
        Logs warnings and errors if the capture device cannot be released or opened.
        Raises:
            Logs exceptions encountered during the initialization process.
        """
        
        if self.cap is not None:
            console_logger.warning("current cap is not None")
            while self.cap is not None and self.cap.isOpened():
                self.cap.release()
                time.sleep(1)
                self.cap = None
            console_logger.info("Release cap successfully")
        try:
            self.source = camera_config_params.camera_source
            if camera_config_params.cap_type_dshow:
                self.cap = cv2.VideoCapture(self.source, cv2.CAP_DSHOW)
                if camera_config_params.enable_mjpg_format:
                    self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter.fourcc(*"MJPG"))
            else:
                self.cap = cv2.VideoCapture(self.source)
                if camera_config_params.enable_mjpg_format:
                    self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter.fourcc(*"MJPG"))

            # self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
            if self.cap is None or not self.cap.isOpened():
                console_logger.error("Cannot opened video capture!")
                return
            if camera_config_params.manual_setting:
                self.brigtness = camera_config_params.brightness
                self.contrast = camera_config_params.contrast
                self.saturation = camera_config_params.saturation
                self.sharpness = camera_config_params.sharpness
                self.gain = camera_config_params.gain
                ####
                self.cap.set(
                    cv2.CAP_PROP_BRIGHTNESS, self.brigtness
                )  # Brightness: thử từ 0 đến 255
                self.cap.set(cv2.CAP_PROP_CONTRAST, self.contrast)
                self.cap.set(cv2.CAP_PROP_SATURATION, self.saturation)
                self.cap.set(cv2.CAP_PROP_SHARPNESS, self.sharpness)
                self.cap.set(cv2.CAP_PROP_GAIN, self.gain)

            self.cap.set(
                cv2.CAP_PROP_FRAME_WIDTH, camera_config_params.camera_resolution[1]
            )
            self.cap.set(
                cv2.CAP_PROP_FRAME_HEIGHT, camera_config_params.camera_resolution[0]
            )
            self.cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, camera_config_params.auto_exposure)
            # 0.25 = Manual mode for some cameras
            if isinstance(camera_config_params.prop_exposure, int):
                self.cap.set(cv2.CAP_PROP_EXPOSURE, camera_config_params.prop_exposure)
                
            print("Contrast:", self.cap.get(cv2.CAP_PROP_CONTRAST))
            print("Saturation:", self.cap.get(cv2.CAP_PROP_SATURATION))
            print("Sharpness:", self.cap.get(cv2.CAP_PROP_SHARPNESS))  # Không phải camera nào cũng hỗ trợ
            print("Gain:", self.cap.get(cv2.CAP_PROP_GAIN))
        except Exception as e:
            console_logger.error(f"Cannot open video capture: {e}\n")
            operation_history_log.error(f"Cannot open video capture: {e}\n")

    def start_stream_video(self) -> None:
        """
        Starts the video streaming process, capturing frames from the configured video source,
        processing them according to the current application state, and emitting signals with
        the processed frames and calculated FPS.
        The method performs the following steps:
        - Initializes frame count and running state.
        - Continuously reads frames from the video capture device while streaming is active.
        - Handles dynamic changes in video source and reinitializes capture if needed.
        - Applies rotation and resizing to frames based on camera configuration parameters.
        - Calculates the frames per second (FPS) for each frame.
        - Depending on the parent controller's state:
            - If not in auto-process mode, optionally calculates and draws the focus score,
              and emits the frame signal.
            - If in auto-process mode, optionally checks frame stability, performs abnormal
              detection (e.g., dust detection), and emits the appropriate frame signal with
              additional data such as heatmaps, segmentation, or abnormality status.
        - Releases the video capture resource and emits a finished signal when streaming ends.
        Emits:
            emit_frame_signal: Signal with processed frame data and FPS (and optionally abnormality info).
            finished_signal: Signal emitted when streaming is finished.
        Note:
            This method is intended to be run in a separate thread or process to avoid blocking the UI.
        """
        
        self.count_fr = 0
        self.is_running = True
        # self.init_video_capture()
        while self.is_running and self.cap is not None and self.cap.isOpened():
            start = time.time()
            if self.source != camera_config_params.camera_source:
                self.init_video_capture()
                continue
            ret, frame = self.cap.read()
            if not ret:
                continue

            if camera_config_params.is_need_rotate_camera:
                frame = self.rotate_image(frame, camera_config_params.rotate_angle)

            if self.source != 0:
                frame = cv2.resize(
                    frame,
                    (
                        camera_config_params.camera_resolution[1],
                        camera_config_params.camera_resolution[0],
                    ),
                )
            if camera_config_params.is_need_resize_frame:
                frame = cv2.resize(
                    frame,
                    (
                        camera_config_params.resize_resolution[1],
                        camera_config_params.resize_resolution[0],
                    ),
                )
            end = time.time()
            fps = 1 / (end - start + 1e-5)
            self.count_fr += 1
            # process at here
            if not self._parent.is_running_stream:
                console_logger.debug("is running stream: false")
                continue
            if not self._parent.is_auto_process:
                if system_config_params.show_image_focus_score:
                    now = datetime.now().second
                    if now % system_config_params.skip_frames_show_focus == 0:
                        self.focus_score = math.floor(calculate_focus_score(frame))
                    draw_focus_score(frame, self.focus_score)

                self.emit_frame_signal.emit([frame, fps])
                self._parent.is_abnormal = None
            else:
                if detect_lcd_params.is_need_check_stable_frames:
                    should_process, score = quick_stability_check(frame)
                    console_logger.debug(
                        f"Quick stability check: should_process={should_process}, score={score}"
                    )
                    if not should_process:
                        self.emit_frame_signal.emit([frame, fps])
                        # self._parent.is_abnormal = None
                        continue

                image_with_heatmap, image_with_dust, is_abnormal, original_box = (
                    None,
                    None,
                    None,
                    None,
                )
                start_time = datetime.now()
                if self.abnormal_processing is not None:
                    abnormal_detect_result = (
                        self.abnormal_processing.dust_detect_on_image(frame)
                    )
                end_time = datetime.now()
                fps = 1 / ((end_time - start_time).total_seconds() + 1e-5)
                if abnormal_detect_result is None:
                    self.emit_frame_signal.emit([frame, fps])
                    self._parent.is_abnormal = None
                    console_logger.warning(
                        "Abnormal detection returned None, skipping frame."
                    )
                else:
                    image_with_heatmap, image_with_dust, is_abnormal, original_box = (
                        abnormal_detect_result
                    )
                    if self._parent.ui_heat_map_radio_btn.isChecked():
                        self.emit_frame_signal.emit(
                            [image_with_heatmap, fps, is_abnormal, original_box]
                        )
                    elif self._parent.ui_segment_radio_btn.isChecked():
                        self.emit_frame_signal.emit(
                            [image_with_dust, fps, is_abnormal, original_box]
                        )
                    else:
                        self.emit_frame_signal.emit(
                            [frame, fps, is_abnormal, original_box]
                        )

        if self.cap is not None:
            self.cap.release()
            self.cap = None
        self.finished_signal.emit()

    def stop(self) -> None:
        """
        Stops the camera capture process, releases the video capture resource if it is open,
        emits a finished signal, and closes any OpenCV windows.
        This method sets the running flag to False, emits a signal indicating that the process
        has finished, and ensures that the video capture object (`self.cap`) is properly released.
        It also destroys all OpenCV windows and logs the successful release of the camera resource.
        """
        
        self.is_running = False
        self.finished_signal.emit()
        if self.cap is not None:
            while self.cap is not None and self.cap.isOpened():
                self.cap.release()
                time.sleep(1)
                self.cap = None
                cv2.destroyAllWindows()
            console_logger.info("Stop camera successfully, release cap done!")

# # project
class SFCWorker(QObject):
    # finished = Signal(object)
    sent_request_data = Signal(object)

    def __init__(self, request_func: Callable) -> None:
        super().__init__()
        self._data: dict | None = None
        self.request_func = request_func

    @Slot(dict)
    def run(self, data: dict) -> None:
        """
        Executes the main process by sending a request with the provided data and emitting the response.
        Args:
            data (dict): The data to be sent in the request.
        Emits:
            sent_request_data: Signal emitted with the response from the request.
        Logs:
            Logs the request data and any exceptions that occur during the request process.
        """
        
        try:
            console_logger.debug(f"Request data: {data}")
            # đây là hàm gọi API
            response = self.request_func(data=data)
            self.sent_request_data.emit(response)

        except Exception as e:
            console_logger.debug("failed exception request SFC: ", e)

class MainWindowController(QMainWindow, Ui_MainWindow):
    send_data_signal: Signal = Signal(dict)

    def __init__(self) -> None:
        super().__init__()
        self.setupUi(self)
        self.init_variables()
        self.mock_funtions_base()
        self.init_style_sheet()
        self.create_sfc_req_res_thread()
        self.init_thread_stream_video()
        operation_history_log.info("Open application\n")
                
        self.filter = ResizeFilter(self.ui_label_display_video)
        self.ui_label_display_video.installEventFilter(self.filter)

    def init_variables(self) -> None:
        """
        Initializes and resets instance variables related to the main window controller's state.
        This method sets up default values for system status, UI controls, process flags, 
        product counters, and threading objects. It also configures the initial state of 
        various UI elements based on current configuration parameters and input values.
        Variables initialized include:
            - System and process status flags (e.g., is_abnormal, is_running_stream)
            - Image and MAC address input fields
            - Debug and auto-process flags from configuration
            - Product OK/NG counters from configuration
            - UI control states (enabled/disabled, checked/unchecked)
            - Video streaming worker and thread placeholders
            - Operation timing and cycle time tracking
        No arguments or return values.
        """
        
        # system
        self.ok_frame_count = 0  # count abnormal frame to decided NG or OK
        self.status_result = None  # status result of process
        self.is_abnormal: None | bool = True  # is abnormal or not
        self.image_need_process = ""  # image in process type is image
        self.is_running_stream = False  # is running stream camera
        self.input_mac_value = ""  # mac address input
        self.is_auto_process = system_config_params.is_auto_process
        self.is_debug = system_config_params.debug
        self.process_type = 0
        self.ui_segment_radio_btn.setChecked(True)
        self.current_processing_image = None
        self.is_calling_sfc = False  # default process type is segment

        self.sum_ok_products = product_config_params.sum_ok_products
        self.sum_ng_products = product_config_params.sum_ng_products
        if self.image_need_process == "":
            self.ui_process_btn.setDisabled(True)
        else:
            self.ui_process_btn.setEnabled(True)
            # self.ui_process_btn.setStyleSheet("border: 2px solid a9dfbf ")
        self.ui_stop_camera_btn.setDisabled(True)
        self.ui_mac_input.setEnabled(False)

        self.ui_auto_process_check_box.setChecked(self.is_auto_process)
        self.ui_debug_check_box.setChecked(self.is_debug)
        self.ui_auto_process_check_box.setEnabled(False)

        self.stream_video_worker = None
        self.stream_video_thread = None

        self.start_operation = 0
        self.end_operation = 0
        self.cycle_time = 0

        self.cycle_time_list = []

    def mock_funtions_base(self) -> None:
        """
        Sets up mock functions and connects UI elements to their respective event handlers.
        This method initializes various UI event handlers for debugging, image processing, 
        camera streaming, configuration management, and input focus. It includes:
            - Processing a single image using the dust detection algorithm and displaying the result.
            - Changing the debug state and updating the logger level.
            - Starting and stopping the camera stream.
            - Opening an image file and updating the UI accordingly.
            - Saving all configuration settings.
            - Handling changes to the auto-process state.
            - Handling changes to the MAC input field.
            - Opening the configuration file.
            - Setting focus to the MAC input field at regular intervals.
        Raises:
            Exception: If an error occurs during image processing, a message box is displayed with the error message.
        """
        

        def process_one_image():
            try:
                if self.stream_video_worker is not None:
                    if self.stream_video_worker.abnormal_processing is not None:
                        result = self.stream_video_worker.abnormal_processing.dust_detect_on_image(
                            self.image_need_process
                        )
                        if result is not None:
                            if not self.ui_heat_map_radio_btn.isChecked():
                                self.ui_label_display_video.setImage(result[1])
                            else:
                                self.ui_label_display_video.setImage(result[0])
                            print("ok")
            except Exception as e:
                QMessageBox.information(None, "Error", f"{e}")

        # change debug console_logger level
        def on_debug_state_changed(state):
            self.is_debug = self.ui_debug_check_box.isChecked()
            print(f"Debug: {self.is_debug}")
            system_config_params.debug = self.is_debug
            console_logger.setLevel(
                logging.DEBUG if self.ui_debug_check_box.isChecked() else logging.INFO
            )

        self.ui_debug_check_box.stateChanged.connect(on_debug_state_changed)

        # start and stop stream camera
        self.ui_start_camera_btn.clicked.connect(self.start_stream)
        self.ui_stop_camera_btn.clicked.connect(self.stop_stream)

        # open image button
        def click_open_image():
            # Mở hộp thoại chọn file ảnh
            file_path, _ = QFileDialog.getOpenFileName(
                None, "Open Image", "", "Image Files (*.png *.jpg *.bmp *.jpeg *.WEBP)"
            )
            if file_path:
                self.ui_process_btn.setDisabled(False)
                self.image_need_process = file_path
                self.update_process_type_mode()
                # self.set_background_label(self.ui_label_display_video, file_path)
                self.ui_label_display_video.setImage(file_path)

            if self.image_need_process != "":
                self.ui_process_btn.setEnabled(True)
                # self.ui_process_btn.setStyleSheet("border: 2px solid a9dfbf ")
            else:
                self.ui_process_btn.setEnabled(True)

        self.ui_open_image_btn.clicked.connect(click_open_image)
        self.ui_process_btn.clicked.connect(process_one_image)

        # save all config button
        self.ui_save_all_config_btn.triggered.connect(lambda: save_all_config())

        # auto process button change state
        def handle_autoprocess_change(state):
            self.is_auto_process = self.ui_auto_process_check_box.isChecked()
            system_config_params.is_auto_process = self.is_auto_process

        self.ui_auto_process_check_box.stateChanged.connect(handle_autoprocess_change)

        # input mac change value

        self.ui_mac_input.textChanged.connect(self.handle_mac_input_changed)

        # open config button
        # set focus to mac input
        self.focus_timer = QTimer(self)
        self.focus_timer.setInterval(system_config_params.time_to_focus_mac_input)
        self.focus_timer.timeout.connect(
            lambda: self.ui_mac_input.setFocus()
            if not self.ui_mac_input.hasFocus()
            else None
        )
        self.focus_timer.start()

    def update_process_type_mode(self) -> None:
        """Updates the visual style of process type labels based on the current processing mode.
        If `self.image_need_process` is an empty string, highlights the camera process type label
        and resets the image process type label's style. Otherwise, highlights the image process type
        label and resets the camera process type label's style.
        This method is typically used to provide visual feedback to the user about the current
        process type selection in the UI."""
        
    def init_style_sheet(self) -> None:
        """
        Initializes and applies the style sheet and UI resources for the main window.
        This method performs the following actions:
            - Updates the process type mode.
            - Sets a custom style sheet for the 'Open Config File' button, including hover effects.
            - Sets background images for specific labels using provided resource paths.
            - Sets the display image for the video label asynchronously.
            - Configures icons and icon sizes for the 'Start Camera' and 'Stop Camera' buttons.
        Returns:
            None
        """
        
        self.update_process_type_mode()

        # self.set_background_label(self.ui_label_display_video,get_resource_path(r"src\assets\IVIS.png"))
        
        QTimer.singleShot(
            0,
            lambda: self.ui_label_display_video.setImage(
                get_resource_path("src/assets/IVIS.png")
            ),
        )

        icon = QIcon()
        icon.addFile(
            get_resource_path(r"src\assets\start.jpg"),
            QSize(),
            QIcon.Mode.Normal,
            QIcon.State.Off,
        )
        self.ui_start_camera_btn.setIcon(icon)
        self.ui_start_camera_btn.setIconSize(QSize(17, 17))

        icon1 = QIcon()
        icon1.addFile(
            get_resource_path(r"src\assets\stop.png"),
            QSize(),
            QIcon.Mode.Normal,
            QIcon.State.Off,
        )
        self.ui_stop_camera_btn.setIcon(icon1)
        self.ui_stop_camera_btn.setIconSize(QSize(27, 27))

    def on_stream_thread_finished(self) -> None:
        """
        Handles UI updates when the stream thread has finished.
        This method resets and updates various UI components to their default or inactive states,
        including enabling/disabling buttons, clearing input fields, updating status and system
        messages, and setting default images and styles. It is typically called when a background
        streaming operation completes, ensuring the UI reflects the idle or ready state.
        """
        
        self.ui_open_image_btn.setDisabled(False)
        self.ui_auto_process_check_box.setEnabled(False)
        self.ui_mac_input.setEnabled(False)
        self.ui_start_camera_btn.setDisabled(False)
        self.ui_stop_camera_btn.setDisabled(True)

        self.ui_mac_input.setText("")
        self.ui_system_message_label.setText("This is system message")
        self.ui_system_message_label.setStyleSheet(
            "background-color: white;font-weight:bold"
        )

        # self.set_background_label(
        #     self.ui_label_display_video, get_resource_path(r"src\assets\IVIS.png")
        # )
        self.ui_label_display_video.setImage(get_resource_path(r"src\assets\IVIS.png"))
        self.ui_status_label.setStyleSheet(
            system_config_params.normal_status_label_style
        )
        self.ui_status_label.setText(system_config_params.normal_status_label_text)

    def init_thread_stream_video(self) -> None:
        """
        Initializes and starts a new QThread for streaming video.
        This method creates a new QThread and a StreamVideoWorker instance, moves the worker to the thread,
        and connects the necessary signals and slots for video streaming and thread management. If a stream
        thread is already running, it stops the current thread before starting a new one. Handles exceptions
        and logs relevant information.
        Raises:
            Exception: If initialization of the stream video thread fails.
        """
        
        if self.stream_video_thread is not None:
            console_logger.warning("Current stream thread is not None")
            self.stop_current_thread()
        try:
            self.stream_video_thread = QThread()
            self.stream_video_worker = StreamVideoWorker(self)
            self.stream_video_worker.moveToThread(self.stream_video_thread)
            self.stream_video_thread.started.connect(
                self.stream_video_worker.start_stream_video
            )
            self.stream_video_worker.emit_frame_signal.connect(
                self.update_stream_camera
            )

            self.stream_video_worker.finished_signal.connect(
                self.stream_video_thread.quit
            )

            self.stream_video_thread.finished.connect(self.on_stream_thread_finished)

            self.stream_video_thread.finished.connect(
                self.stream_video_worker.deleteLater
            )
            self.stream_video_thread.finished.connect(
                self.stream_video_thread.deleteLater
            )

            # self.normal_stream_video_worker.blockSignals(False)
            self.stream_video_thread.start()

            console_logger.info("Init stream video thread successfully!")
        except Exception as e:
            console_logger.error(f"Failed when init stream video thread: {e}")

    def stop_current_thread(self) -> None:
        """
        Stops the currently running video streaming thread safely.
        This method attempts to stop the associated worker if it exists,
        waits for the thread to finish within a timeout (5 seconds),
        and forcefully terminates the thread if it does not stop in time.
        Logs appropriate messages for successful or unsuccessful termination.
        Resets the thread and worker references to None after stopping.
        Raises:
            RuntimeError: If an error occurs while stopping the thread.
        """
        
        if self.stream_video_thread is not None:
            try:
                if self.stream_video_worker is not None:
                    self.stream_video_worker.stop()
                if not self.stream_video_thread.wait(5000):
                    console_logger.warning("Thread không dừng được, force terminate")
                    self.stream_video_thread.terminate()
                    self.stream_video_thread.wait()

                console_logger.info(
                    "Thread đã dừng thành công"
                )  # Wait for the thread to finish
            except RuntimeError as e:
                print(f"Error stopping thread: {e}")

            self.stream_video_thread = None
            self.stream_video_worker = None
            print(1)

    def start_stream(self) -> None:
        """
        Starts the camera stream and updates the UI accordingly.
        This method performs the following actions:
            - Clears the image to be processed.
            - Disables the 'Start Camera' button to prevent multiple starts.
            - Logs the start of the camera operation.
            - Sets the internal flag indicating the stream is running.
            - Enables or disables relevant UI elements for camera operation.
        """
        
        # check if is exist stream camera thread
        self.image_need_process = ""
        self.ui_start_camera_btn.setDisabled(True)
        operation_history_log.info("Start camera!\n")
        console_logger.info("Start camera!")
        self.is_running_stream = True
        # self.init_thread_stream_video()
        # change UI
        self.ui_auto_process_check_box.setEnabled(True)
        self.ui_mac_input.setEnabled(True)
        self.ui_open_image_btn.setDisabled(True)
        self.ui_stop_camera_btn.setDisabled(False)

    def stop_stream(self) -> None:
        self.ui_stop_camera_btn.setDisabled(True)
        operation_history_log.info("Stop camera!\n")
        console_logger.info("Stop camera!")
        self.is_calling_sfc = False
        self.is_running_stream = False
        self.input_mac_value = ""
        self.input_tray_value = ""
        # self.set_background_label(self.ui_label_display_video,get_resource_path(r'src\assets\IVIS.png'))
        self.ui_label_display_video.setImage(get_resource_path(r"src\assets\IVIS.png"))
        self.ui_start_camera_btn.setDisabled(False)
        self.ui_mac_input.setDisabled(True)
        self.ui_open_image_btn.setDisabled(False)
        # self.stop_current_thread()

    def set_background_label(self, label: QLabel, cv_image: np.ndarray | str | None) -> None:
        """
        Sets the background image of a QLabel widget using an OpenCV image or image file path.
        This method converts the provided image (either as a NumPy array or a file path) to RGB format,
        resizes it to fit the QLabel while maintaining aspect ratio, and sets it as the label's pixmap.
        If `cv_image` is None, the method does nothing.
        Args:
            label (QLabel): The QLabel widget to set the background image for.
            cv_image (np.ndarray | str | None): The image to display. Can be a NumPy array (OpenCV image),
                a file path to an image, or None.
        Returns:
            None
        """
        
        if cv_image is None:
            return
        if isinstance(cv_image, str):
            cv_image = cv2.imread(cv_image)
        # Chuyển sang RGB (OpenCV default là BGR)
        rgb_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w

        label_size = label.size()
        label_width = label_size.width()
        label_height = label_size.height()

        q_img = QImage(
            rgb_image.data, w, h, bytes_per_line, QImage.Format.Format_RGB888
        )
        pixmap = QPixmap.fromImage(q_img).scaled(
            label_width,
            label_height,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        # Căn giữa ảnh trong label
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setPixmap(pixmap)

    def expand_box(self, box, expand_pixels):
        """
        Expands a given quadrilateral box by a specified number of pixels in all directions.
        This function calculates the minimum area rotated rectangle that encloses the input box,
        expands its width and height by `2 * expand_pixels`, and returns the coordinates of the
        expanded rectangle as a polygon.
        Args:
            box (array-like): An array of shape (4, 2) representing the four corner points of the box.
            expand_pixels (int or float): The number of pixels to expand the box in all directions.
        Returns:
            numpy.ndarray: An array of shape (4, 1, 2) containing the integer coordinates of the expanded box corners.
        """
        # Bước 1: Tính rotated rectangle (minAreaRect)
        rect = cv2.minAreaRect(box)  # rect = ((cx, cy), (w, h), angle)

        # Bước 2: Mở rộng kích thước
        (cx, cy), (w, h), angle = rect
        w_exp = w + 2 * expand_pixels
        h_exp = h + 2 * expand_pixels
        rect_expanded = ((cx, cy), (w_exp, h_exp), angle)

        # Bước 3: Lấy lại polygon từ rect mở rộng
        box_points = cv2.boxPoints(rect_expanded)  # shape (4, 2)
        box_points = np.int32(box_points).reshape(-1, 1, 2)
        return box_points

    @Slot(object)
    def update_stream_camera(
        self,
        data: list[np.ndarray | int] | tuple[np.ndarray, int, bool, list[int]]
    ):
        
        """
        Updates the stream camera display and UI status based on the provided data.
        This method processes incoming frame data from the camera, updates the display image,
        manages abnormal/normal status detection, and updates UI labels and styles accordingly.
        Args:
            data (list[np.ndarray | int] | tuple[np.ndarray, int, bool, list[int]]):
                - If length is 2: [display_image, fps]
                    - display_image (np.ndarray): The image frame to display.
                    - fps (int): Frames per second value.
                - If length is 4: [display_image, fps, is_abnormal, original_box]
                    - display_image (np.ndarray): The image frame to display.
                    - fps (int): Frames per second value.
                    - is_abnormal (bool): Whether the current frame is considered abnormal.
                    - original_box (list[int]): Coordinates for drawing a bounding box.
        Side Effects:
            - Updates the UI status label's text and style based on the detection result.
            - Draws bounding boxes and FPS text on the display image.
            - Updates the displayed image in the UI.
            - Maintains internal state for abnormal/normal frame counting.
        Note:
            This method assumes the existence of several instance attributes and external
            configuration parameters, such as `self.ui_status_label`, `self.ui_label_display_video`,
            `system_config_params`, `abnormal_inference_params`, and `detect_lcd_params`.
        """
        
        # if not self.is_auto_process:
        if len(data) == 2:
            self.ok_frame_count = 0
            # self.is_abnormal = True
            display_image, fps = data
            self.ui_status_label.setStyleSheet(
                system_config_params.normal_status_label_style
            )

            if self.is_abnormal is None:
                self.ui_status_label.setText(
                    system_config_params.normal_status_label_text
                )
                self.ui_status_label.setStyleSheet(
                    system_config_params.normal_status_label_style
                )

        else:
            display_image, fps, is_abnormal, original_box = data
            if detect_lcd_params.is_need_scan_sn_first:
                if original_box is not None and self.input_mac_value == "":
                    # self.set_system_message(f"{sfc_request_params.scan_sn_first_message}", 'red')
                    self.ui_status_label.setText("SN???")
                    self.ui_status_label.setStyleSheet(
                        system_config_params.ng_status_label_style
                    )
                    # self.set_background_label(self.ui_label_display_video, display_image)
                    self.ui_label_display_video.setImage(display_image)
                    return
            if not is_abnormal:
                self.ok_frame_count += 1
                if (
                    self.ok_frame_count
                    >= abnormal_inference_params.frame_count_threashold_ok
                ):
                    self.is_abnormal = False
                    self.ok_frame_count = 0
            else:
                self.ok_frame_count = 0
                self.is_abnormal = True

            if self.is_abnormal and original_box is not None:
                display_image = cv2.polylines(
                    display_image, # type: ignore
                    [original_box], # type: ignore
                    True,
                    abnormal_inference_params.ng_color,
                    2,
                    cv2.LINE_AA,
                )  # type: ignore
            elif not self.is_abnormal and original_box is not None:
                display_image = cv2.polylines(
                    display_image, # type: ignore
                    [original_box], # type: ignore
                    True,
                    abnormal_inference_params.ok_color,
                    2,
                    cv2.LINE_AA,
                )  # type: ignore

            cv2.putText(
                display_image,  # type: ignore
                f"FPS: {fps:.1f}",
                (10, 30),
                fontFace=cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=1,
                color=(0, 255, 0),
                thickness=2,
            )  # type: ignore

            if display_image is not None:
                self.current_processing_image = display_image.copy()  # type: ignore # Lưu ảnh hiện tại để xử lý sau

            if not self.is_abnormal:
                self.ui_status_label.setStyleSheet(
                    system_config_params.ok_status_label_style
                )
                self.ui_status_label.setText(system_config_params.ok_status_label_text)

            elif self.is_abnormal:
                self.ui_status_label.setStyleSheet(
                    system_config_params.ng_status_label_style
                )
                self.ui_status_label.setText(system_config_params.ng_status_label_text)
        # self.set_background_label(self.ui_label_display_video, display_image)
        self.ui_label_display_video.setImage(display_image)

    def set_system_message(
        self,
        text: str,
        color: Literal["red", "green"] = "red"
    ) -> None:
        self.ui_system_message_label.setText(text)
        self.ui_system_message_label.setStyleSheet(
            f"background-color: {color}; font-weight: bold; color: white"
        )

    def validator_mac_input(self, mac_value: str) -> bool:
        return mac_value.startswith(product_config_params.mac_start_with)

    def handle_mac_input_changed(self, new_text: str) -> None:
        min_len = product_config_params.mac_min_len
        if (
            len(new_text) >= min_len
            and not self.is_calling_sfc
            and self.is_running_stream
            and self.is_auto_process
        ):
            is_valid_mac = self.validator_mac_input(new_text[0:min_len])
            if is_valid_mac:
                self.input_mac_value = new_text[0:min_len]
                self.ui_mac_input.setText("")
                self.set_system_message(
                    sfc_request_params.notify_text_valid_sn, color="green"
                )
                self.start_operation = time.time()
                operation_history_log.info(f"Scan SN: {self.input_mac_value} - PASS")

                return
            if self.input_mac_value == "":
                self.set_system_message(
                    sfc_request_params.notify_text_invalid_sn, color="red"
                )
                operation_history_log.warning(f"SN scan failed: {self.input_mac_value}")
                self.ui_mac_input.setText("")
            else:
                if len(new_text) >= product_config_params.tray_min_len:
                    if self.is_abnormal is False:
                        self.set_system_message(
                            sfc_request_params.notify_text_valid_tray, color="green"
                        )
                        operation_history_log.info(
                            f"SN: {self.input_mac_value} - scan TRAY PASS"
                        )
                        self.input_tray_value = new_text[
                            0 : product_config_params.tray_min_len
                        ]
                        data_sfc = sfc_request_params.Data_send_to_SFC
                        keys = list(data_sfc.keys())
                        data_sfc[keys[0]] = sfc_request_params.LINE_NAME
                        data_sfc[keys[1]] = sfc_request_params.GROUP_NAME
                        data_sfc[keys[2]] = sfc_request_params.SP
                        data_sfc[keys[3]] = (
                            f"SN={self.input_mac_value},TRAY={self.input_tray_value},EMP={sfc_request_params.EMP},PASSED={sfc_request_params.PASSED}"
                        )
                        # self.is_calling_sfc = True
                        self.send_sfc_request(data_sfc)
                    else:
                        operation_history_log.info(
                            f"SN: {self.input_mac_value} - scan TRAY failed - dusting"
                        )
                        self.set_system_message(
                            sfc_request_params.notify_text_still_dusty, "red"
                        )
                    self.ui_mac_input.setText("")

    def send_sfc_request(self, data: dict) -> None:
        if self.is_calling_sfc:
            return  # tránh xử lý đồng thời

        self.is_calling_sfc = True

        # Gọi method process của worker trong thread
        self.send_data_signal.emit(data)

    def create_sfc_req_res_thread(self) -> None:
        # self.is_calling_sfc = True
        try:
            self.sfc_request_thread = QThread()
            self.sfc_worker = SFCWorker(self.request_sfc)
            self.sfc_worker.moveToThread(self.sfc_request_thread)
            self.send_data_signal.connect(self.sfc_worker.run)

            # self.sfc_request_thread.started.connect(self.sfc_worker.run)
            self.sfc_worker.sent_request_data.connect(self.handle_sfc_result)
            self.sfc_request_thread.start()
            console_logger.info("Create sfc request thread successfully!")
        except Exception as e:
            console_logger.error(f"Failed when create sfc request thread: {e}")

    def handle_sfc_result(self, sfc_response: dict) -> None:
        self.is_calling_sfc = False
        if sfc_response is not None:
            try:
                if (
                    not isinstance(sfc_response, dict)
                    or "Data" not in sfc_response
                    or "RES" not in sfc_response["Data"]
                ):
                    operation_history_log.info(f"SFC Falied: {sfc_response}\n\n")
                    raise ValueError("cannot call SFC")
                if (
                    sfc_response[sfc_request_params.sfc_response_key1][
                        sfc_request_params.sfc_response_key2
                    ]
                    == sfc_request_params.SFC_Response_status
                ):
                    try:
                        self.ui_mac_input.setText("")
                        self.sum_ok_products += 1
                        product_config_params.sum_ok_products += 1
                        product_config_params.save_to_config_file()
                        self.ui_sum_ok_label.setText(f"N OK: {self.sum_ok_products}")
                        self.set_system_message(
                            sfc_request_params.notify_text_passed_product, color="green"
                        )
                        self.save_pass_image(
                            self.current_processing_image, self.input_mac_value
                        )
                        self.end_operation = time.time()
                        self.cycle_time = self.end_operation - self.start_operation
                        self.cycle_time_list.append(self.cycle_time)
                        operation_history_log.info(
                            f"Save image successfully - MAC: {self.input_mac_value} - PASS\nCycle time: {self.cycle_time}\n\n"
                        )
                        self.write_log_to_excel(self.input_mac_value)

                    except Exception as e:
                        console_logger.error(f"Error : {e}")
                        operation_history_log.info(
                            f"SFC Falied: {self.input_mac_value} - Mess: {e}\n\n"
                        )
                else:
                    console_logger.debug(f"SFC response: {sfc_response}")
                    self.set_system_message(
                        f"Command1: {sfc_response[sfc_request_params.sfc_response_key1][sfc_request_params.sfc_response_key3]}",
                        color="red",
                    )
                    operation_history_log.info(
                        f"SFC failed: {self.input_mac_value} \n\n "
                    )
            except (KeyError, ValueError, TypeError) as e:
                self.set_system_message(
                    f"{sfc_request_params.notify_text_sfc_failed_request}: {str(e)}",
                    color="red",
                )
            finally:
                self.input_mac_value = ""
                self.input_tray_value = ""

    def request_sfc(
        self,
        url: str = sfc_request_params.end_point,
        method: str = sfc_request_params.method,
        data: None | dict = None,
        headers: dict = sfc_request_params.headers,
    ) -> dict | str:
        self.sfc_failed = True
        try:
            if method == "GET":
                operation_history_log.info(f"Request SFC data: {data}")
                response = requests.get(url, headers=headers)

            elif method == "POST":
                operation_history_log.info(f"Request SFC data: {data}")
                response = requests.post(url, json=data, headers=headers)

            operation_history_log.info(f"SFC response: {response.json()}")
            if response.status_code == 200:
                return response.json()
            else:
                res = f"Lỗi http {response.status_code}: {response.text}"

                return res
        except Exception as e:
            res = f"Lỗi khi gọi API: {e}"
            return res

    def write_log_to_excel(
        self, log_text: str, log_dir: str = sfc_request_params.default_dir_log_sn_pass
    ) -> None:
        # Đảm bảo thư mục log tồn tại
        os.makedirs(log_dir, exist_ok=True)

        # Tạo tên file dạng dd-mm-yyyy.xlsx
        file_name = datetime.now().strftime("%d-%m-%Y") + ".xlsx"
        file_path = os.path.join(log_dir, file_name)

        # Lấy thời gian hiện tại (hh:mm:ss)
        current_time = datetime.now().strftime("%H:%M:%S")

        # Nếu file đã tồn tại, mở và ghi tiếp, ngược lại tạo mới
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Scan Time", "SN", "Cycle time"])  # type: ignore # Tiêu đề cột

        # Ghi thêm dòng log
        ws.append([current_time, log_text, self.cycle_time])  # type: ignore
        wb.save(file_path)

    def save_pass_image(
        self,
        image: np.ndarray | None,
        text: str,
        base_dir: str = sfc_request_params.default_dir_log_image,
    ) -> None:
        # Lấy thời gian hiện tại
        if image is None:
            console_logger.error("Image is none, cannot save")
            operation_history_log.error("Image is none, cannot save")
            return
        now = datetime.now()
        timestamp_str = now.strftime("%d/%m/%Y %H:%M:%S")  # Dùng để vẽ lên ảnh
        folder_name = now.strftime("%d-%m-%Y")  # Dùng để tạo thư mục
        time_str = now.strftime("%H_%M_%S")  # Dùng cho tên file

        # Tạo thư mục theo ngày nếu chưa có
        full_dir = os.path.join(base_dir, folder_name)
        os.makedirs(full_dir, exist_ok=True)

        # Copy ảnh để không sửa ảnh gốc
        img = image.copy()

        # Cấu hình font
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = sfc_request_params.font_scale_log_image
        thickness = sfc_request_params.thickness_text_log_image
        color = sfc_request_params.color_text_log_image  # Đỏ

        # Tính kích thước text
        (w1, h1), _ = cv2.getTextSize(timestamp_str, font, font_scale, thickness)
        (w2, h2), _ = cv2.getTextSize(f"MAC: {text}", font, font_scale, thickness)

        # Vị trí text: góc dưới bên phải
        h, w = img.shape[:2]
        margin = 10
        x = w - max(w1, w2) - margin
        y1 = h - margin - h2
        y2 = h - margin

        # Vẽ thời gian và MAC lên ảnh
        cv2.putText(
            img,
            timestamp_str,
            (x, y1 - 20),
            font,
            font_scale,
            color,
            thickness,
            cv2.LINE_AA,
        )
        cv2.putText(
            img,
            f"MAC: {text}",
            (x, y2),
            font,
            font_scale,
            color,
            thickness,
            cv2.LINE_AA,
        )

        # Tạo tên file
        filename = os.path.join(
            full_dir, f"{time_str}_MAC_{text}_{'NG' if self.is_abnormal else 'OK'}.jpg"
        )

        # Lưu ảnh
        cv2.imwrite(filename, img)
        console_logger.info(f"Save image {filename} successfully")
        print(f"✅ Ảnh đã lưu: {filename}")

    def closeEvent(self, event) -> None:
        self.stop_current_thread()
        if self.sfc_worker is not None:
            self.sfc_worker.deleteLater()
        if self.sfc_request_thread is not None:
            self.sfc_request_thread.quit()
            self.sfc_request_thread.wait()
            self.sfc_request_thread.deleteLater()

        operation_history_log.info("Close application\n")
        print("Closing application...")

if __name__ == "__main__":
    app = QApplication([])
    window = MainWindowController()
    window.show()
    app.exec()
